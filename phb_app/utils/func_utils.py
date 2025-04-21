'''
Package
-------
Utilities

Module Name
---------
Employee Management Utilities

Version
-------
Date-based Version: 20250304
Author: Karl Goran Antony Zuvela

Description
-----------
Provides utility functions for the management
of employees in the project hours budgeting wizard.
'''

import zipfile
from typing import TYPE_CHECKING, Iterator
from datetime import datetime
from functools import lru_cache
import xlwings as xw
from PyQt6.QtCore import QModelIndex
from PyQt6.QtWidgets import QTableWidget
from openpyxl import load_workbook
import openpyxl.utils as xlutils
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.exceptions import (
    ReadOnlyWorkbookException,
    InvalidFileException
)
from phb_app.logging.exceptions import (
    CountryIdentifiersNotInFilename,
    WorkbookLoadError
)
from phb_app.logging.exceptions import (
    MissingEmployeeRow,
    EmployeeRowAnchorsMisalignment
)

if TYPE_CHECKING:
    from phb_app.data.phb_dataclasses import (
        CountryData,
        EmployeeRange,
        EmployeeRowAnchors,
        ManagedWorkbook,
        ManagedOutputWorkbook,
        ManagedInputWorkbook,
        CountriesEnum,
        ProjectIDTableHeaders,
        Employee
    )

#####################
### File handling ###
#####################

def get_origin_from_file_name(file_name: str, country_data: "CountryData", countries_enum: "CountriesEnum") -> str:
    '''
    Checks the workbook's origin. Returns the country of origin.
    '''
    # Pattern in all German timesheets
    german_patterns = country_data.get_locale_by_country(countries_enum.GERMANY.value).file_patterns
    # Pattern in all Romanian timesheets
    romanian_patterns = country_data.get_locale_by_country(countries_enum.ROMANIA.value).file_patterns
    if all(pattern in file_name.lower() for pattern in german_patterns):
        return countries_enum.GERMANY.value
    if all(pattern in file_name.lower() for pattern in romanian_patterns):
        return countries_enum.ROMANIA.value
    raise CountryIdentifiersNotInFilename(file_name)

def is_workbook_open(file_path:str) -> bool:
    '''Check if the file is already in use.'''
    try:
        with open(file_path, 'r+', encoding='utf-8'):
            pass
    except (OSError, IOError):
        return True
    return False

def try_load_workbook(workbook: "ManagedWorkbook", wb_type: "ManagedOutputWorkbook"):
    '''
    Template for attempting to load the workbook.
    '''
    file_name = workbook.file_name
    file_path = workbook.file_path
    try:
        if isinstance(workbook, wb_type):
            # We only care about the output workbook being open
            # as we will not write to the input workbook
            with open(file_path, 'r+', encoding='utf-8'):
                pass
        workbook = load_workbook(file_path)
        return workbook
    except ReadOnlyWorkbookException as e:
        raise WorkbookLoadError(f"Workbook '{file_name}' is read-only: {str(e)}.") from e
    except InvalidFileException as e:
        raise WorkbookLoadError(f"Invalid Excel file '{file_name}': {str(e)}.") from e
    except FileNotFoundError as e:
        raise WorkbookLoadError(f"File '{file_name}' not found.") from e
    except PermissionError as e:
        raise WorkbookLoadError(f"Permission denied: Close '{file_name}' if open.") from e
    except zipfile.BadZipFile as e:
        raise WorkbookLoadError(f"Corrupted Excel file '{file_name}': {str(e)}") from e

def is_file_already_in_table(file_path: str, col: int, *tables: QTableWidget) -> bool:
    '''
    Check if the file already exists in either input or output table
    and return a respective boolean.
    '''
    return any(
        table.item(row, col)
        and table.item(row, col).text() == file_path
        for table in tables
        for row in range(table.rowCount())
    )

def replace_in_string(string: str, old: str, new: str) -> str:
    '''
    Replaces the old with new in the given string.
    '''
    return string.replace(old, new)

#####################
### Date handling ###
#####################

def german_abbr_month(month_number: int, months: dict) -> str:
    '''
    Change a month number to its short name.
    '''
    for month_key, month_value in months.items():
        if int(month_value) == int(month_number):
            return month_key
    return ""

# Cache the function results
# Only cache one result to minimise memory use
@lru_cache(maxsize=1)
def get_budgeting_dates(file_path: str, sheet_name: str) -> list[tuple[int, int, int]]:
    '''
    Returns a list of tuples, each tuple containing the month, year and coordinate
    retreived from the possible budgeting dates in the budgeting file.
    '''
    # Do not diplay Excel while computing
    app = xw.App(visible=False)
    wb = app.books.open(file_path)
    sheet = wb.sheets[sheet_name]
    # 52 was chosen because the dates begin there in every file.
    # This could be automated to find yet another dubious entry
    # and the key for that entry could be kept in a yaml config
    # for the user to define. Note: expand stops at the first
    # empty cell.
    col1 = sheet.range('A52').expand('down')
    months_years_rows = [(cell.value.month, cell.value.year, cell.row)
                         for cell in col1
                         if isinstance(cell.value, datetime)]
    wb.close()
    # Excel will quit if no other workbooks are open
    return months_years_rows

#########################
### Employee handling ###
#########################

# Cache the function results
# Only cache one result to minimise memory use
@lru_cache(maxsize=1)
def locate_employee_range(sheet_obj: Worksheet, emp_range: "EmployeeRange", anchors: "EmployeeRowAnchors") -> None:
    '''
    Finds the row range where the employee names should be located.
    '''
    # Temp variables to hold the cell data
    start_anchor_temp = None
    end_anchor_temp = None
    # Loop through every row
    for row in sheet_obj.iter_rows():
        # Loop through every cell of that row
        for cell in row:
            if cell.value == anchors.start_anchor:
                start_anchor_temp = cell
            elif cell.value == anchors.end_anchor:
                end_anchor_temp = cell
    if start_anchor_temp and end_anchor_temp:
        if start_anchor_temp.row == end_anchor_temp.row:
            emp_range.start_cell = start_anchor_temp.coordinate
            emp_range.end_cell = end_anchor_temp.coordinate
        else:
            raise EmployeeRowAnchorsMisalignment(anchors.start_anchor, anchors.end_anchor)
    else:
        if not start_anchor_temp and not end_anchor_temp:
            raise MissingEmployeeRow(anchors.start_anchor, anchors.end_anchor)
        if not start_anchor_temp and end_anchor_temp:
            raise MissingEmployeeRow(anchors.start_anchor)
        if start_anchor_temp and not end_anchor_temp:
            raise MissingEmployeeRow(anchors.end_anchor)

def yield_hours_coord(coord: str, row: int) -> Iterator[str]:
    '''
    Yields the coordinate for where the hours are located per employee
    for the given date (row).
    '''
    # The first item ([0] -> col) in the tuple from `coordinate_from_string` is used
    yield f"{str(xlutils.cell.coordinate_from_string(coord)[0])}{str(row)}"

def find_predicted_hours(emp_dict: dict[str, "Employee"], row: int, file_path: str, sheet_name: str) -> dict[str, int]:
    '''
    Goes through all given coordinates of a worksheet, computes
    any formulae and returns the hours by employee name coordinate.
    '''
    # Do not diplay Excel while computing
    app = xw.App(visible=False)
    wb = app.books.open(file_path)
    sheet = wb.sheets[sheet_name]
    # Prepare a dictionary of coord:hours
    pre_hours = {}
    for emp_coord, emp in emp_dict.items():
        # Create a coordinate from the date's row and employee's column
        hours_coord = next(yield_hours_coord(emp_coord, row))
        # Save the computed value
        pre_hours[emp_coord] = sheet.range(hours_coord).value
        # Save the hours coordinate
        emp.hours.hours_coord = hours_coord
    wb.close()
    return pre_hours

########################
### Project handling ###
########################

def set_selected_project_ids(mngd_wb: "ManagedInputWorkbook", table: QTableWidget, rows: list[QModelIndex], headers: "ProjectIDTableHeaders") -> None:
    '''
    Sets the selected projects IDs as references from the selectable IDs.
    '''
    for row in rows:
        # Get the project ID and file name from each row
        proj_id = table.item(row.row(), headers.PROJECT_ID.value).text()
        file_name = table.item(row.row(), headers.FILENAME.value).text()
        if file_name == mngd_wb.file_name:
            # If the file name of the managed input workbook is correct,
            # make a reference by key from selected to selectable project IDs
            mngd_wb.managed_sheet_object.selected_project_ids[proj_id] = mngd_wb.managed_sheet_object.selectable_project_ids[proj_id]
