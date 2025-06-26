'''
Package
-------
File Handling Utilities

Module Name
---------
Employee Management Utilities

Author
-------
Karl Goran Antony Zuvela

Description
-----------
Provides file handling utility functions in the project hours budgeting wizard.
'''

import zipfile
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.utils.exceptions import (
    ReadOnlyWorkbookException,
    InvalidFileException
)
from PyQt6.QtWidgets import QTableWidget
import phb_app.data.location_management as loc
import phb_app.logging.exceptions as ex
import phb_app.wizard.constants.ui_strings as st
import phb_app.templating.types as t

def get_origin_from_file_name(
    file_name: str,
    country_data: loc.CountryData,
    countries_enum: st.CountriesEnum
) -> t.CountryName:
    '''    Checks the workbook's origin. Returns the country of origin.'''
    file_name_lower = file_name.lower()
    for country in countries_enum:
        locale = country_data.get_locale_by_country(country)
        # Assumes every locale has a .file_patterns attribute
        patterns = locale.file_patterns
        if all(pattern in file_name_lower for pattern in patterns):
            return country
    raise ex.CountryIdentifiersNotInFilename(file_name)

def is_workbook_open(file_path:str) -> bool:
    '''Check if the file is already in use.'''
    try:
        with open(file_path, 'r+', encoding='utf-8'):
            pass
    except (OSError, IOError):
        return True
    return False

def try_load_workbook(file_path: str, file_name: str, writable: bool = False) -> Workbook:
    '''Template for attempting to load the workbook.'''
    try:
        if writable:
            # We only care about the output workbook, which is writable
            # being open as we will not write to the input workbook
            with open(file_path, 'r+', encoding='utf-8'):
                pass
        return load_workbook(file_path)
    except ReadOnlyWorkbookException as e:
        raise ex.WorkbookLoadError(f"Workbook '{file_name}' is read-only: {str(e)}.") from e
    except InvalidFileException as e:
        raise ex.WorkbookLoadError(f"Invalid Excel file '{file_name}': {str(e)}.") from e
    except FileNotFoundError as e:
        raise ex.WorkbookLoadError(f"File '{file_name}' not found.") from e
    except PermissionError as e:
        raise ex.WorkbookLoadError(f"Permission denied: Close '{file_name}' if open.") from e
    except zipfile.BadZipFile as e:
        raise ex.WorkbookLoadError(f"Corrupted Excel file '{file_name}': {str(e)}") from e

def is_file_already_in_table(file_path: str, col: int, table: QTableWidget) -> bool:
    '''Check if there are two or more of the same file in the given table and return
    a respective boolean.'''
    return 1 < sum(
        1 for row in range(table.rowCount())
        if table.item(row, col) and table.item(row, col).text() == file_path
        )
