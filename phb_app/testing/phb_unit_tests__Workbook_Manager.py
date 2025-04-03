import unittest
from AutoHoursCollector import WorkbookManager
from openpyxl import Workbook
from unittest.mock import patch, MagicMock, PropertyMock
from os import path

############################################################
# Class to test the workbook manager.
class TestWorkbookManager(unittest.TestCase):

    def create_patch(self, target_name, is_class=False, object_name=None, static_method=False, autospec=True):
        if not static_method and not is_class:
            patcher = patch(target_name, autospec=autospec)
            patched_function = patcher.start()
            self.addCleanup(patcher.stop)
            return patched_function
        elif static_method and not is_class:
            patcher = patch.object(object_name, target_name, autospec=autospec)
            patched_object = patcher.start()
            self.addCleanup(patcher.stop)
            return patched_object
        elif is_class:
            patcher = patch(target_name)
            patched_class = patcher.start()
            self.addCleanup(patcher.stop)
            return patched_class

    def setUp(self):
        """Test prerequisites."""
        # Patch objects and functions
        self.create_patch('builtins.print') # Suppress prints statements
        self.mock_askopenfilename = self.create_patch('tkinter.filedialog.askopenfilename')
        self.mock_load_workbook = self.create_patch('AutoHoursCollector.load_workbook')
        # Common return values
        self.mock_askopenfilename.return_value = 'c:/test.xlsx'
        self.test_file_name = path.basename(self.mock_askopenfilename.return_value)
        self.mock_load_workbook.return_value = MagicMock(name="Workbook")
        self.test_sheet_name = "Tabelle1"

    ############################################################
    # Test cases for the class method open_new_workbook.
    
    @patch.object(WorkbookManager, 'is_workbook_open', return_value=False, autospec=True)
    def test_successfully_opening_new_workbook(self, mock_is_workbook_open):
        """Test the class method to open new workbooks with a valid file path."""
        mock_worksheet_manager_instance = self.create_patch('AutoHoursCollector.WorksheetManager', is_class=True).return_value
        type(mock_worksheet_manager_instance).select_sheet_name = PropertyMock()
        mock_worksheet_manager_instance.selected_sheet = self.test_sheet_name
        mock_worksheet_manager_instance.select_worksheet_gui.side_effect = lambda: None # GUI triggered. No real return required.
        workbook_manager_instance = WorkbookManager.open_new_workbook(title="Select an Excel file", is_pivot_table=True) # Method under test
        # Assertions for the real object
        self.assertIsNotNone(workbook_manager_instance)
        self.assertEqual(workbook_manager_instance.worksheet_manager.selected_sheet, self.test_sheet_name)
        self.assertEqual(workbook_manager_instance.file_name, self.test_file_name)
        self.assertTrue(workbook_manager_instance.is_pivot_table)
        self.assertIsNotNone(workbook_manager_instance.worksheet_manager)
        # Assertions for mocks
        mock_is_workbook_open.assert_called_with(self.mock_askopenfilename.return_value)
        self.mock_load_workbook.assert_called_once_with(self.mock_askopenfilename.return_value, data_only=True)
        self.assertIsInstance(workbook_manager_instance.worksheet_manager, MagicMock)
        mock_worksheet_manager_instance.select_worksheet_gui.assert_called_once()

    @patch('tkinter.messagebox.showerror')
    @patch.object(WorkbookManager, 'is_workbook_open', return_value=True, autospec=True)
    def test_opening_already_open_workbook(self, mock_is_workbook_open, mock_showerror):
        """Test the class method to open new workbooks with a valid file path but already open workbook.
        The workbook <file name> is in use. Please save and close it before continuing."""
        workbook_manager_instance = WorkbookManager.open_new_workbook(title="Select an Excel file", is_pivot_table=True) # Method under test
        # Assertions for the real object
        self.assertIsNone(workbook_manager_instance)
        # Assertions for mocks
        mock_is_workbook_open.assert_called_with(self.mock_askopenfilename.return_value)
        mock_showerror.assert_called_with("Error", f"The workbook {self.test_file_name} is in use.\nPlease save and close it before continuing.")

    @patch("sys.exit")
    @patch.object(WorkbookManager, 'is_workbook_open', return_value=False, autospec=True)
    def test_no_worksheet_selected(self, mock_is_workbook_open,mock_sys_exit):
        """Test the class method to open new workbooks with a valid file path but
        no worksheet selected from the new workbook."""
        mock_worksheet_manager_instance = self.create_patch('AutoHoursCollector.WorksheetManager', is_class=True).return_value
        type(mock_worksheet_manager_instance).select_sheet_name = PropertyMock()
        mock_worksheet_manager_instance.selected_sheet = None
        mock_worksheet_manager_instance.select_worksheet_gui.side_effect = lambda: None # GUI triggered. No real return required.
        workbook_manager_instance = WorkbookManager.open_new_workbook(title="Select an Excel file", is_pivot_table=True) # Method under test
        # Assertions for the real object
        self.assertIsNone(workbook_manager_instance)
        # Assertions for mocks
        mock_is_workbook_open.assert_called_with(self.mock_askopenfilename.return_value)
        self.mock_load_workbook.assert_called_once_with(self.mock_askopenfilename.return_value, data_only=True)
        mock_worksheet_manager_instance.select_worksheet_gui.assert_called_once()
        mock_sys_exit.assert_called_once()

    @patch("sys.exit")
    @patch.object(WorkbookManager, 'is_workbook_open', return_value=False, autospec=True)
    def test_no_filename_opening_new_workbook(self, mock_is_workbook_open, mock_sys_exit):
        """Test the class method to open new workbooks with no valid file path,
        i.e. \"No new workbook selected.\""""
        self.mock_askopenfilename.return_value = None
        self.test_file_name = None
        workbook_manager_instance = WorkbookManager.open_new_workbook(title="Select an Excel file", is_pivot_table=True) # Method under test
        # Assertions for the real object
        self.assertIsNone(workbook_manager_instance)
        mock_sys_exit.assert_called_once()

    ############################################################
    # Test cases for the static method is_workbook_open.

    def test_is_workbook_open_yes_oserror(self):
        """Test the static method to check if the workbook is already open (OSError)."""
        with patch("builtins.open", side_effect=OSError):
            result = WorkbookManager.is_workbook_open(self.mock_askopenfilename.return_value) # Method under test
        # Assertions for the real object
        self.assertTrue(result)
    
    def test_is_workbook_open_yes_ioerror(self):
        """Test the static method to check if the workbook is not already open (IOError)."""
        with patch("builtins.open", side_effect=IOError):
            result = WorkbookManager.is_workbook_open(self.mock_askopenfilename.return_value) # Method under test
        # Assertions for the real object
        self.assertTrue(result)

    def test_is_workbook_open_no(self):
        """Test the static method to check if the workbook is not already open (IOError)."""
        with patch("builtins.open"):
            result = WorkbookManager.is_workbook_open(self.mock_askopenfilename.return_value) # Method under test
        # Assertions for the real object
        self.assertFalse(result)

############################################################
    # Test cases for the function init_excel_workbook.

    def test_init_excel_workbook_error(self):
        """Test the function to initialise the workbook but an error occurs."""
        self.mock_load_workbook.side_effect = Exception("Failed to load workbook.")
        workbook_manager_instance = WorkbookManager()
        workbook_manager_instance.init_excel_workbook(self.mock_askopenfilename.return_value)
        # Assertions for the real object
        self.assertIsNone(workbook_manager_instance.workbook_object)
        self.assertEqual(workbook_manager_instance.open_workbooks, [])

    @patch.object(WorkbookManager, 'is_workbook_open', return_value=False, autospec=True)
    def test_init_excel_workbook_success(self, mock_is_workbook_open):
        """Test the function to initialise the workbook."""
        workbook_manager_instance = WorkbookManager()
        workbook_manager_instance.init_excel_workbook(self.mock_askopenfilename.return_value)
        # Assertions for the real object
        self.assertIsNotNone(workbook_manager_instance.workbook_object)
        self.assertIn(workbook_manager_instance, WorkbookManager.open_workbooks)

    ############################################################
    # Test cases for the function get_file_name. Getting the filename successfully is already tested
    # indirectly in TCs for open_new_workbook.

    def test_get_file_name_unnamed(self):
        """Test the function to get the file name but the file has no name."""
        workbook_manager_instance = WorkbookManager()
        workbook_manager_instance.file_path = None
        workbook_manager_instance.get_file_name()
        # Assertions for the real object
        self.assertEqual(workbook_manager_instance.file_name, "Unnamed workbook")
    
    ############################################################
    # Test cases for the function extract_year_month_from_target_file_name.

    def test_extract_year_month_from_target_file_name_full_no_match(self):
        """Test the function to get the date from the file name but no date is found."""
        workbook_manager_instance = WorkbookManager()
        workbook_manager_instance.file_name = "test.xlsx"
        workbook_manager_instance.extract_year_month_from_target_file_name()
        # Assertions for the real object
        self.assertEqual(workbook_manager_instance.extract_year_month_from_target_file_name(), (None, None))

    def test_extract_year_month_from_target_file_name_partial_match(self):
        """Test the function to get the date from the file name but no full expected date is found."""
        workbook_manager_instance = WorkbookManager()
        workbook_manager_instance.file_name = "test_2024_0_test.xlsx"
        workbook_manager_instance.extract_year_month_from_target_file_name()
        # Assertions for the real object
        self.assertEqual(workbook_manager_instance.extract_year_month_from_target_file_name(), (None, None))

    def test_extract_year_month_from_target_file_name_match_found(self):
        """Test the function to get the date from the file name successfully."""
        workbook_manager_instance = WorkbookManager()
        workbook_manager_instance.file_name = "test_2024_11_test.xlsx"
        workbook_manager_instance.extract_year_month_from_target_file_name()
        # Assertions for the real object
        self.assertEqual(workbook_manager_instance.extract_year_month_from_target_file_name(), ("2024", "11"))

    ############################################################
    # Clean up.

    def tearDown(self):
        """Clean up any persistent state after each test."""
        WorkbookManager.open_workbooks.clear()
        patch.stopall()
    
############################################################
# Main test block.
if __name__ == '__main__':
    unittest.main()