import unittest
import pyautogui as pag
import threading
import os
import tempfile
import pygetwindow as gw
from AutoHoursCollector import WorkbookManager, WorksheetManager
from openpyxl import Workbook, load_workbook
from unittest.mock import patch, MagicMock, PropertyMock
from datetime import datetime
from io import BytesIO
import tkinter as tk
from PyQt5.QtWidgets import QApplication, QPushButton, QMessageBox
from PyQt5.QtTest import QTest
from PyQt5.QtCore import Qt, QTimer

unittest.TestLoader.sortTestMethodsUsing = None

############################################################
# Class to test the worksheet manager.
class TestWorksheetManager(unittest.TestCase):

    # Test data variables
    MONTH_MAPPING = {
        "Jan": 1, "Feb": 2, "Mrz": 3, "Apr": 4, "Mai": 5, "Jun": 6,
        "Jul": 7, "Aug": 8, "Sep": 9, "Okt": 10, "Nov": 11, "Dez": 12
    }

    pivot_table_headers_list = ["PSP-Element", "Auftrag", "Objektbezeichnung", "Name der P", "Bezeichnung", "BuchDatum", "Menge gesamt", "MEH"]
    pivot_table_headers_dict = {key: value for value, key in enumerate(pivot_table_headers_list)}
    pivot_table_psp_elements = ["PSP1", "PSP2", "PSP3", "PSP4", "PSP5", "PSP7"]

    def create_mock_budget_excel_file(start_cell, end_cell):
        wb = Workbook()
        ws = wb.active
        ws.title = "Tabelle1"
        # The test data runs from column A to I
        test_data = [[start_cell, "Herbert Hübler", "Jasmin Hindersmann", "Birgit Soost", "", "", "Yannick Finke", "", end_cell],
                     ["2024"],
                     ["Jan 2024", "1", "1", "2", "10"],
                     ["Feb 2024", "1", "1", "", "10"],
                     ["Mrz 2024"],
                     ["01"],
                     ["2001"],
                     [""],
                     ["20001"],
                     ["01"],
                     ["20q24"],
                     ["2025x"]]
        # Add rows of data to test whether the correct years are extracted from the column
        for row_data in test_data:
            ws.append(row_data)
        mock_file = BytesIO()
        wb.save(mock_file)
        mock_file.seek(0) # Reset the position to the beginning
        return mock_file
    
    def create_mock_pivot_excel_file():
        wb = Workbook()
        ws = wb.active
        ws.title = "Tabelle1"
        ws.append(TestWorksheetManager.pivot_table_headers_list)
        test_row_data = TestWorksheetManager.pivot_table_psp_elements
        for i, value in enumerate(test_row_data, start=2): # Start below the headers
            ws.cell(row=i, column=1, value=value)
        mock_file = BytesIO()
        wb.save(mock_file)
        mock_file.seek(0) # Reset the position to the beginning
        return mock_file
    
    def open_mock_file(mocked_file):
        """Open the mocked file in Excel to check that the data is suitably mocked.
        create_mock_budget_excel_file, create_mock_pivot_excel_file or an inline created Excel
        file should have already been created before using this function."""
        if mocked_file:
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as temp_file:
                mocked_file.save(temp_file.name)
                os.startfile(temp_file.name)

    def calculate_test_month(self, num_test_steps, mock_file_name_month):
        """Calculates the test month. If the calculated month number exceeds 12 (December), December is returned.
        Number of test steps is how many times \"down\" or \"up\" was used:how many times going down or up
        in the month dropdown."""
        for k, v in self.MONTH_MAPPING.items():
                if int(v) == int(mock_file_name_month):
                    offset = int(v)
                    month_calc = num_test_steps + offset
                    if 0 < month_calc and month_calc < len(self.MONTH_MAPPING):
                        for k, v in self.MONTH_MAPPING.items():
                            if int(v) == month_calc:
                                return k
                    elif month_calc >= len(self.MONTH_MAPPING):
                        return list(self.MONTH_MAPPING.keys())[-1]
                    elif month_calc <= 0:
                        return list(self.MONTH_MAPPING.keys())[0]
    
    def calculate_test_year(self, num_test_steps, mock_file_name_year, mock_year_list):
        """Calculates the test month. If the calculated month number exceeds the year list length,
        the last item of the list is returned."""
        # test_steps is from the test steps: how many times going down in the month dropdown
        year_calc = num_test_steps + int(mock_file_name_year)
        if int(min(mock_year_list)) < year_calc and year_calc < int(max(mock_year_list)):
            return str(year_calc)
        elif year_calc >= int(max(mock_year_list)):
            return mock_year_list[-1]
        elif year_calc <= int(min(mock_year_list)):
            return mock_year_list[0]
    
    def activate_window(expected_title):
        expected_window = gw.getWindowsWithTitle(expected_title)
        if expected_window:
            expected_window[0].activate()
            return True
        return False
    
    def simulate_actions(expected_title, test_steps):
            """Performs the test steps using pyautogui.
            Half a second is given for the GUI to be created.
            Each step is followed by a hardcoded 100ms sleep.
            Parameters:
                expected_title (str): The title of the window to be activated.
                test_steps (list[str]): A list of key presses.
            Returns:
                None"""
            pag.sleep(0.5)
            if TestWorksheetManager.activate_window(expected_title):
                for step in test_steps:
                    if expected_title == gw.getActiveWindowTitle():
                        pag.press(step)
                        pag.sleep(0.1)

    def create_patch(self, target_name, is_class=False, object_name=None, static_method=False, autospec=True):
        if not static_method and not is_class:
            patcher = patch(target_name, autospec=autospec)
            patched_function = patcher.start()
            self.addCleanup(patcher.stop)
            return patched_function
        elif static_method and not is_class:
            patcher = patch.object(object_name, target_name, autospec=autospec)
            patched_object = patcher.start()
            self.addCleanup(patcher.stop)
            return patched_object
        elif is_class:
            patcher = patch(target_name)
            patched_class = patcher.start()
            self.addCleanup(patcher.stop)
            return patched_class

    def setUp(self):
        """Test prerequisites."""
        # Patch objects and functions
        #self.create_patch('builtins.print') # Suppress prints statements
        self.mock_sys_exit = self.create_patch('sys.exit')
        #self.mock_tk_error_warning = self.create_patch('tkinter.messagebox.showerror')
        self.mock_workbook_manager = MagicMock()
        self.mock_workbook_manager.file_name = "temp_PS18_2024_09_Pivot_final.xlsx"
        self.mock_employee_list = ["MA Name\nStartdatum", "Herbert Hübler", "Jasmin Hindersmann", "Birgit Soost", "Yannick Finke"]
        # Common return values
        type(self.mock_workbook_manager).workbook_object = PropertyMock()
        self.mock_workbook_manager.workbook_object.sheetnames = ["Tabelle1", "Tabelle2"]
        
        # Create the class under test
        self.worksheet_manager = WorksheetManager(self.mock_workbook_manager)
        self.app = QApplication([])

    ############################################################
    # Test cases begin

    ######## Get sheet names ########

    def test_get_sheet_names__successful(self):
        """Test the class function to get sheet names successfully."""
        self.worksheet_manager.get_sheet_names()
        self.assertEqual(self.worksheet_manager.sheet_names, self.mock_workbook_manager.workbook_object.sheetnames)
    
    def test_get_sheet_names__fail(self):
        """Test the class function to get sheet names unsuccessfully."""
        self.mock_workbook_manager.workbook_object.sheetnames = None
        self.worksheet_manager.get_sheet_names()
        self.assertIsNone(self.worksheet_manager.sheet_names)
    
    # ####### Select worksheet GUI ########

    def test_select_worksheet_gui__successful_confirm(self):
        """Test the class function select worksheet GUI to successfully select a worksheet."""
        expected_title = "Selectable worksheets"
        # Test steps: Select the first sheet, then go to and click confirm
        test_steps = ['tab', 'down', 'tab', 'space']
        action_thread = threading.Thread(target=TestWorksheetManager.simulate_actions, args=(expected_title, test_steps,), daemon=True) # Function under test in its own thread
        action_thread.start()
        # Test steps
        self.worksheet_manager.select_worksheet_gui()
        self.assertEqual(self.worksheet_manager.selected_sheet['sheet_name'], "Tabelle1")
        action_thread.join()

    @patch('tkinter.messagebox.showerror')
    def test_select_worksheet_gui__empty_confirm_and_cancel(self, mock_showerror):
        """Test the class function select worksheet GUI  does not allow an empty selection to be confirmed and cancelling works."""
        expected_title = "Selectable worksheets"
        # Test steps: Skip selection and click confirm. Return to selection, choose something then go to cancel and click
        test_steps = ['tab', 'tab', 'space', 'space', 'tab', 'tab', 'down', 'tab', 'tab', 'space']
        action_thread = threading.Thread(target=TestWorksheetManager.simulate_actions, args=(expected_title, test_steps,), daemon=True) # Function under test in its own thread
        action_thread.start()
        # Test steps
        self.worksheet_manager.select_worksheet_gui()
        mock_showerror.assert_called_with("Input Error", "Please select a sheet.")
        self.assertEqual(self.worksheet_manager.selected_sheet['sheet_name'], "")
        self.mock_sys_exit.assert_called_once()
        pag.sleep(0.5)
        action_thread.join()

    # ####### Select PSP element GUI ########

    # def test_select_psp_element_gui__not_pivot_table(self):
    #     """Test the GUI class function to select the PSP element but the file is not a pivot table."""
    #     self.mock_workbook_manager.is_pivot_table = False
    #     self.worksheet_manager.select_psp_element_gui()
    #     self.assertIsNone(self.worksheet_manager.selected_psp_element)

    # def test_select_psp_element_gui__is_pivot_table_no_psp(self):
    #     """Test the GUI class function to select the PSP element but the file is a pivot table but with no PSP element."""
    #     mock_get_psp_element = self.create_patch('AutoHoursCollector.WorksheetManager.get_psp_elements')
    #     mock_get_psp_element.return_value = None
    #     self.mock_workbook_manager.is_pivot_table = True
    #     self.worksheet_manager.select_psp_element_gui()
    #     self.assertIsNone(self.worksheet_manager.selected_psp_element)
    
    # def test_select_psp_element_gui__confirm_success(self):
    #     """Test the GUI class function to select the PSP element, successfully confirming the selection."""
    #     expected_title = "Select PSP Element"
    #     # Test steps:
    #     test_steps = ['tab', 'down', 'down', 'down', 'tab', 'space']
    #     mock_get_psp_element = self.create_patch('AutoHoursCollector.WorksheetManager.get_psp_elements')
    #     mock_get_psp_element.return_value = ["PSP1", "PSP2", "PSP3", "PSP4"]
    #     action_thread = threading.Thread(target=TestWorksheetManager.simulate_actions, args=(expected_title, test_steps,), daemon=True) # Function under test in its own thread
    #     action_thread.start()
    #     # Test GUI
    #     self.worksheet_manager.select_psp_element_gui()    
    #     self.assertEqual(self.worksheet_manager.selected_psp_element, "PSP3")
        # action_thread.join()
    
    # @patch('tkinter.messagebox.showerror')
    # def test_select_psp_element_gui__empty_selection_and_cancel(self, mock_showerror):
    #     """Test the GUI class function to select the PSP element. The selection is empty
    #         and the user cancels the operation."""
    #     expected_title = "Select PSP Element"
    #     # Test steps: Go to and click confirm, go to the PSP dropdown, select first, go to and click confirm
    #     test_steps = ['tab', 'tab', 'space', 'tab', 'tab', 'down', 'tab', 'tab', 'space']
    #     mock_get_psp_element = self.create_patch('AutoHoursCollector.WorksheetManager.get_psp_elements')
    #     mock_get_psp_element.return_value = ["PSP1", "PSP2", "PSP3", "PSP4"]
    #     action_thread = threading.Thread(target=TestWorksheetManager.simulate_actions, args=(expected_title, test_steps,), daemon=True) # Function under test in its own thread
    #     action_thread.start()
    #     # Test GUI
    #     self.worksheet_manager.select_psp_element_gui()
    #     mock_showerror.assert_called_with("Input Error", "Please select a PSP element.")
    #     self.assertIsNone(self.worksheet_manager.selected_psp_element)
    #     self.mock_sys_exit.assert_called_once()
    #     action_thread.join()
     
    ####### Select month-year GUI ########

    # def test_select_month_year_gui__successful_confirm(self):
    #     """Test the class function select month and year GUI to successfully select a month and year.
    #     The default month and year shall be taken from the file name."""
    #     expected_title = "Choose the month and year of budgeted hours"
    #     # Test steps: select month dropdown, go down twice, select year dropdown, go down three times, select and click confirm
    #     test_steps = ['tab', 'down', 'down', 'tab', 'down', 'down', 'down', 'tab', 'space']
    #     month_steps_down = 1
    #     year_steps_down = 2
    #     mock_extract_budgeting_years = self.create_patch('AutoHoursCollector.WorksheetManager.extract_budgeting_years') # Skip functionality. Not needed further
    #     self.worksheet_manager.budgeting_years = ["2021", "2022", "2023", "2024"]
    #     mock_file_name_year = "2021"
    #     mock_file_name_month = "05"
    #     self.mock_workbook_manager.extract_year_month_from_target_file_name.return_value = (mock_file_name_year, mock_file_name_month)
    #     #action_thread = threading.Thread(target=self.worksheet_manager.select_month_year_gui, args=(self.mock_workbook_manager,), daemon=True) # Function under test in its own thread
    #     action_thread = threading.Thread(target=TestWorksheetManager.simulate_actions, args=(expected_title, test_steps,), daemon=True) # Function under test in its own thread
    #     action_thread.start()
    #     # Test GUI
    #     self.worksheet_manager.select_month_year_gui(self.mock_workbook_manager)
    #     self.assertEqual(self.worksheet_manager.selected_month, self.calculate_test_month(month_steps_down, mock_file_name_month))
    #     self.assertEqual(self.worksheet_manager.selected_year, self.calculate_test_year(year_steps_down, mock_file_name_year, self.worksheet_manager.budgeting_years))
    #     action_thread.join()

    # def test_select_month_year_gui__no_date_filename_and_confirm(self):
    #     """Test the class function select month and year GUI. The month and year will be empty.
    #     The default month and year is not found in the file name."""
    #     mock_extract_budgeting_years = self.create_patch('AutoHoursCollector.WorksheetManager.extract_budgeting_years') # Skip functionality. Not needed further
    #     self.worksheet_manager.budgeting_years = [str(datetime.now().year + offset) for offset in range(-5,1)]
    #     self.mock_workbook_manager.extract_year_month_from_target_file_name.return_value = (None, None)
    #     mock_file_name_year = datetime.now().year # See whether the current year is used.
    #     mock_file_name_month = datetime.now().month # See whether the current month is used.
    #     action_thread = threading.Thread(target=self.worksheet_manager.select_month_year_gui, args=(self.mock_workbook_manager,), daemon=True) # Function under test in its own thread
    #     action_thread.start()
    #     # Test steps
    #     pag.sleep(0.05)
    #     pag.press('tab') # Select dropdown
    #     pag.sleep(0.1)
    #     pag.press('down') # Pull down month dropdown menu
    #     pag.sleep(0.1)
    #     pag.press('down') # Go down one month from default
    #     pag.sleep(0.1)
    #     months_steps = 1
    #     pag.press('tab') # go to year dropdown
    #     pag.sleep(0.1)
    #     pag.press('down') # Pull down year dropdown menu
    #     pag.sleep(0.1)
    #     pag.press('down') # Go down one year from default
    #     pag.sleep(0.1)
    #     pag.press('down') # Go down another year
    #     pag.sleep(0.1)
    #     year_steps = 2
    #     pag.press('tab') # go to confirm button
    #     pag.sleep(0.1)
    #     pag.press('space') # Click confirm button
    #     pag.sleep(0.1)
    #     self.assertEqual(self.worksheet_manager.selected_month, self.calculate_test_month(months_steps, mock_file_name_month))
    #     self.assertEqual(self.worksheet_manager.selected_year, self.calculate_test_year(year_steps, mock_file_name_year, self.worksheet_manager.budgeting_years))
    #     pag.sleep(0.5)
    #     action_thread.join()

    # def test_select_month_year_gui__no_date_filename_and_cancel(self):
    #     """Test the class function select month and year GUI. The month and year will be empty.
    #     The default month and year is not found in the file name. Operation cancelled."""
    #     mock_extract_budgeting_years = self.create_patch('AutoHoursCollector.WorksheetManager.extract_budgeting_years') # Skip functionality. Not needed further
    #     self.worksheet_manager.budgeting_years = [str(datetime.now().year + offset) for offset in range(-5,1)]
    #     self.mock_workbook_manager.extract_year_month_from_target_file_name.return_value = (None, None)
    #     mock_file_name_year = datetime.now().year # See whether the current year is used.
    #     mock_file_name_month = datetime.now().month # See whether the current month is used.
    #     action_thread = threading.Thread(target=self.worksheet_manager.select_month_year_gui, args=(self.mock_workbook_manager,), daemon=True) # Function under test in its own thread
    #     action_thread.start()
    #     # Test steps
    #     pag.sleep(0.05)
    #     pag.press('tab') # Select dropdown
    #     pag.sleep(0.1)
    #     pag.press('down') # Pull down month dropdown menu
    #     pag.sleep(0.1)
    #     pag.press('down') # Go down one month from default
    #     pag.sleep(0.1)
    #     months_steps = 1
    #     pag.press('tab') # go to year dropdown
    #     pag.sleep(0.1)
    #     pag.press('down') # Pull down year dropdown menu
    #     pag.sleep(0.1)
    #     pag.press('down') # Go down one year from default
    #     pag.sleep(0.1)
    #     pag.press('down') # Go down another year
    #     pag.sleep(0.1)
    #     year_steps = 2
    #     pag.press('tab') # go to confirm button
    #     pag.sleep(0.1)
    #     pag.press('tab') # go to cancel button
    #     pag.sleep(0.1) 
    #     pag.press('space') # Click cancel button
    #     pag.sleep(0.1)       
    #     self.assertIsNone(self.worksheet_manager.selected_month)
    #     self.assertIsNone(self.worksheet_manager.selected_year)
    #     self.mock_sys_exit.assert_called_once()
    #     pag.sleep(0.5)
    #     action_thread.join()
    
    # ##### Select employees GUI ########

    # def test_select_employees_gui__target_row_not_found(self):
    #     """Test the class function select employees GUI. The target row is not found."""
    #     mock_locate_target_row_employee_names = self.create_patch('AutoHoursCollector.WorksheetManager.locate_target_row_employee_names') # Skip functionality. Not needed further
    #     mock_locate_target_row_employee_names.return_value = False
    #     test_result = [] # Mutable container shared between threads.
    #     def target_wrapper():
    #         test_result.append(self.worksheet_manager.select_employees_gui())
    #     action_thread = threading.Thread(target=target_wrapper, daemon=True) # Function under test in its own thread
    #     action_thread.start() 
    #     # Test steps
    #     pag.sleep(0.05)
    #     TestWorksheetManager.find_tk_message_click_ok(tk._default_root)
    #     pag.sleep(0.1)
    #     action_thread.join()
    #     # Assertions     
    #     self.assertEqual(test_result[0], False)
    #     pag.sleep(0.5)
    
    # def test_select_employees_gui__no_employees_selected_then_selected(self):
    #     """Test the class function select employees GUI. The target row is found but no employees are 
    #     at first selected but then after the error notification, a selection is made."""
    #     # Local test function
    #     def click_qwarning_ok():
    #         active_modal = self.worksheet_manager.window.findChild(QMessageBox)
    #         self.assertIsNotNone(QMessageBox.warning) # QMessage warning exists
    #         ok_button = active_modal.button(QMessageBox.Ok)
    #         QTest.mouseClick(ok_button, Qt.LeftButton)
    #     # Mock functions
    #     mock_locate_target_row_employee_names = self.create_patch('AutoHoursCollector.WorksheetManager.locate_target_row_employee_names') # Skip functionality. Not needed further
    #     mock_locate_target_row_employee_names.return_value = True
    #     # Mock attributes
    #     self.worksheet_manager.employee_range = {'start_cell': 'B1', 'end_cell': 'I1'}
    #     simulated_excel_file = TestWorksheetManager.create_mock_budget_excel_file("MA Name\nStartdatum", "Anzahl\nMA")
    #     mock_openpyxl_workbook = load_workbook(simulated_excel_file)
    #     mock_worksheet_object = mock_openpyxl_workbook["Tabelle1"]
    #     self.worksheet_manager.selected_sheet = {'sheet_name': "Tabelle1", 'sheet_object': mock_worksheet_object}
    #     self.worksheet_manager.select_employees_gui()
    #     # Test steps
    #     QTest.qWait(100) # ms
    #     test_confirm_button = self.worksheet_manager.window.findChild(QPushButton, "confirmButton")
    #     QTimer.singleShot(500, click_qwarning_ok) # Thread a wait for the modal window to activate and be closed
    #     QTest.mouseClick(test_confirm_button, Qt.LeftButton)
    #     QTest.qWait(100) # ms
    #     self.worksheet_manager.employee_list.item(0).setSelected(True) # Herbert Hübler
    #     self.worksheet_manager.employee_list.item(1).setSelected(True) # Jasmin Hindersmann
    #     self.worksheet_manager.employee_list.item(3).setSelected(True) # Yannick Finke
    #     QTest.qWait(100) # ms
    #     QTest.mouseClick(test_confirm_button, Qt.LeftButton)
    #     QTest.qWait(100)
    #     # Assertions     
    #     self.assertEqual(self.worksheet_manager.selected_employees, {"Herbert Hübler": "B1", "Jasmin Hindersmann": "C1", "Yannick Finke": "G1"})

    # def test_select_employees_gui__cancel(self):
    #     """Test the class function select employees GUI. The target row is found but the user cancels the operation."""
    #     # Mock functions
    #     mock_locate_target_row_employee_names = self.create_patch('AutoHoursCollector.WorksheetManager.locate_target_row_employee_names') # Skip functionality. Not needed further
    #     mock_locate_target_row_employee_names.return_value = True
    #     # Mock attributes
    #     self.worksheet_manager.employee_range = {'start_cell': 'B1', 'end_cell': 'I1'}
    #     simulated_excel_file = TestWorksheetManager.create_mock_budget_excel_file("MA Name\nStartdatum", "Anzahl\nMA")
    #     mock_openpyxl_workbook = load_workbook(simulated_excel_file)
    #     mock_worksheet_object = mock_openpyxl_workbook["Tabelle1"]
    #     self.worksheet_manager.selected_sheet = {'sheet_name': "Tabelle1", 'sheet_object': mock_worksheet_object}
    #     self.worksheet_manager.select_employees_gui()
    #     # Test steps
    #     QTest.qWait(100) # ms
    #     test_cancel_button = self.worksheet_manager.window.findChild(QPushButton, "cancelButton")
    #     QTest.mouseClick(test_cancel_button, Qt.LeftButton)
    #     QTest.qWait(100)
    #     # Assertions     
    #     self.assertEqual(self.worksheet_manager.selected_employees, {})
    #     self.mock_sys_exit.assert_called_once()
    
    # ##### Get PSP Elements ########

    # def test_get_psp_elements__not_pivot_table(self):
    #     """ Test the class function which gets the PSP elements from the pivot table.
    #     The selected worksheet is not from the pivot table. Return an empty list."""
    #     self.mock_workbook_manager.is_pivot_table = False
    #     self.assertEqual(self.worksheet_manager.get_psp_elements(), [])
    
    # def test_get_psp_elements__no_selected_sheet(self):
    #     """ Test the class function which gets the PSP elements from the pivot table.
    #     No selected worksheet. Return an empty list."""
    #     self.mock_workbook_manager.is_pivot_table = True
    #     self.worksheet_manager.selected_sheet['sheet_object'] = None
    #     self.assertEqual(self.worksheet_manager.get_psp_elements(), [])
    
    # def test_get_psp_elements__return_sorted_psp_values(self):
    #     """ Test the class function which gets the PSP elements from the pivot table.
    #     The function returns a list of sorted PSP values.
    #     The test case also checks whether empty cells are skipped."""
    #     # Mocking
    #     simulated_excel_file = TestWorksheetManager.create_mock_pivot_excel_file()
    #     mock_openpyxl_workbook = load_workbook(simulated_excel_file)
    #     mock_worksheet_object = mock_openpyxl_workbook["Tabelle1"]
    #     self.worksheet_manager.selected_sheet = {'sheet_name': "Tabelle1", 'sheet_object': mock_worksheet_object}
    #     expected_psp_data = sorted(set(TestWorksheetManager.pivot_table_psp_elements))
    #     # Test steps
    #     self.mock_workbook_manager.is_pivot_table = True
    #     self.worksheet_manager.selected_sheet['sheet_object'] = mock_worksheet_object
    #     self.assertEqual(self.worksheet_manager.get_psp_elements(), expected_psp_data)
    
    # ###### Extract budgeting years ########

    # def test_extract_budgeting_years(self):
    #     """Test the class function which extracts the budgeting years from the hours budgeting file.
    #     The returned sorted list should contain only years with the format YYYY."""
    #     # Mocking
    #     simulated_excel_file = TestWorksheetManager.create_mock_budget_excel_file("MA Name\nStartdatum", "Anzahl\nMA")
    #     mock_openpyxl_workbook = load_workbook(simulated_excel_file)
    #     mock_worksheet_object = mock_openpyxl_workbook["Tabelle1"]
    #     self.worksheet_manager.selected_sheet = {'sheet_name': "Tabelle1", 'sheet_object': mock_worksheet_object}
    #     # Test steps
    #     self.worksheet_manager.extract_budgeting_years() # Function under test
    #     self.assertEqual(self.worksheet_manager.budgeting_years, ["2001", "2024"])
    
    #  ###### Locate target row of employees in budgeting file ########

    # def test_locate_target_row_employee_names__start_term_not_found(self):
    #     """Test the class function to locate the row in the budgeting file containing the
    #     employee names. The first term \"MA Name Startdatum\" is not found. No coordinates saved."""
    #     # Mocking
    #     simulated_excel_file = TestWorksheetManager.create_mock_budget_excel_file("", "Anzahl\nMA")
    #     mock_openpyxl_workbook = load_workbook(simulated_excel_file)
    #     mock_worksheet_object = mock_openpyxl_workbook["Tabelle1"]
    #     self.worksheet_manager.selected_sheet = {'sheet_name': "Tabelle1", 'sheet_object': mock_worksheet_object}
    #     # Test steps
    #     self.assertEqual(self.worksheet_manager.locate_target_row_employee_names(), False) # Function under test
    #     # The test data runs from column A to I
    #     self.assertEqual(self.worksheet_manager.employee_range['start_cell'], None)
    #     self.assertEqual(self.worksheet_manager.employee_range['end_cell'], None)
    
    # def test_locate_target_row_employee_names__end_term_not_found(self):
    #     """Test the class function to locate the row in the budgeting file containing the
    #     employee names. The first term \"Anzahl MA\" is not found. Only the start coordinate is saved."""
    #     # Mocking
    #     simulated_excel_file = TestWorksheetManager.create_mock_budget_excel_file("MA Name\nStartdatum", "")
    #     mock_openpyxl_workbook = load_workbook(simulated_excel_file)
    #     mock_worksheet_object = mock_openpyxl_workbook["Tabelle1"]
    #     self.worksheet_manager.selected_sheet = {'sheet_name': "Tabelle1", 'sheet_object': mock_worksheet_object}
    #     # Test steps
    #     self.assertEqual(self.worksheet_manager.locate_target_row_employee_names(), False) # Function under test
    #     # The test data runs from column A to I
    #     self.assertEqual(self.worksheet_manager.employee_range['start_cell'], 'A1')
    #     self.assertEqual(self.worksheet_manager.employee_range['end_cell'], None)
    
    # def test_locate_target_row_employee_names__employees_found(self):
    #     """Test the class function to locate the row in the budgeting file containing the
    #     employee names. The list of employees is found. The start and end coordinates are saved."""
    #     # Mocking
    #     simulated_excel_file = TestWorksheetManager.create_mock_budget_excel_file("MA Name\nStartdatum", "Anzahl\nMA")
    #     mock_openpyxl_workbook = load_workbook(simulated_excel_file)
    #     mock_worksheet_object = mock_openpyxl_workbook["Tabelle1"]
    #     self.worksheet_manager.selected_sheet = {'sheet_name': "Tabelle1", 'sheet_object': mock_worksheet_object}
    #     # Test steps
    #     self.assertEqual(self.worksheet_manager.locate_target_row_employee_names(), True) # Function under test
    #     # The test data runs from column A to I
    #     self.assertEqual(self.worksheet_manager.employee_range['start_cell'], 'A1')
    #     self.assertEqual(self.worksheet_manager.employee_range['end_cell'], 'I1')
    
    # ####### Index pivot table column headers ########

    # def test_index_pivot_table_column_headers__no_worksheet_selected(self):
    #     """Test the class function to index column headers but no worrksheet was provided."""
    #     self.worksheet_manager.selected_sheet = {'sheet_name': None, 'sheet_object': None}
    #     self.worksheet_manager.index_pivot_table_column_headers() # Function under test
    #     self.assertEqual(self.worksheet_manager.column_headers, {})
    
    # def test_index_pivot_table_column_headers__no_worksheet_selected(self):
    #     """Test the class function to index column headers. The indexing is successful."""
    #     simulated_excel_file = TestWorksheetManager.create_mock_pivot_excel_file()
    #     mock_openpyxl_workbook = load_workbook(simulated_excel_file)
    #     #TestWorksheetManager.open_mock_file(mock_openpyxl_workbook)
    #     mock_worksheet_object = mock_openpyxl_workbook["Tabelle1"]
    #     self.worksheet_manager.selected_sheet = {'sheet_name': "Tabelle1", 'sheet_object': mock_worksheet_object}
    #     self.worksheet_manager.index_pivot_table_column_headers() # Function under test
    #     self.assertEqual(self.worksheet_manager.column_headers, TestWorksheetManager.pivot_table_headers_dict)


    ############################################################
    # Clean up.

    def tearDown(self):
        """Clean up any persistent state after each test."""
        WorkbookManager.open_workbooks.clear()
        patch.stopall()
    
############################################################
# Main test block.
if __name__ == '__main__':
    unittest.main()