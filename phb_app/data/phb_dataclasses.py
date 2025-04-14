'''
Module Name
---------
Project Hours Budgeting Data Classes (phb_dataclasses)

Version
-------
Date-based Version: 20250304
Author: Karl Goran Antony Zuvela

Description
-----------
Provides several data classes for the Project Hours Budgeting Wizard.
'''

from os import path
from datetime import datetime
from enum import Enum, auto
from abc import ABC, abstractmethod
from typing import Optional, Iterator
from dataclasses import dataclass, field
from PyQt6.QtCore import Qt
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
import openpyxl.utils as xlutils
import yaml
from phb_app.logging.exceptions import (
    BudgetingDatesNotFound,
    WorkbookAlreadyTracked
)
import phb_app.utils.func_utils as futils

######################################
#### Localised Months dictionary  ####
######################################

MONATE_KURZ_DE = {
  "Jan": 1,
  "Feb": 2,
  "Mrz": 3,
  "Apr": 4,
  "Mai": 5,
  "Jun": 6,
  "Jul": 7,
  "Aug": 8,
  "Sep": 9,
  "Okt": 10,
  "Nov": 11,
  "Dez": 12
}

##########################
#### Enum base class  ####
##########################

class BaseTableHeaders(Enum):
    '''Base enum class with overridden string dunder method
    and general class method.'''

    @classmethod
    def cap_members_list(cls) -> list:
        '''Return a list of capitalised enum members.'''

        return [name.replace('_', ' ').title() for name in cls.__members__]

    @classmethod
    def list_all_values(cls) -> list[int | str]:
        '''Returns a list of all member values.'''

        return [member.value for member in cls]
    
    def __str__(self):
        return self.name.replace('_', ' ').title()

class WizardPageIDs(BaseTableHeaders):
    '''Enumerated pages in order of appearance.'''

    EXPLANATION_PAGE = 0
    I_O_SELECTION_PAGE = auto()
    PROJECT_SELECTION_PAGE = auto()
    EMPLOYEE_SELECTION_PAGE = auto()
    SUMMARY_PAGE = auto()

##########################
#### IO Selection Enum ####
##########################

class IOTable(Enum):
    '''Input or output.'''

    INPUT = "input"
    OUTPUT = "output"

class QPropName(Enum):
    '''QWizard property names.'''

    MANAGED_WORKBOOKS = "managed_workbooks_property"
    SELECTED_PROJECTS = "selected_projects"
    INPUT_W_FORMULAE = "input_with_formulae"
    INPUT_WO_FORMULAE = "input_without_formulae"
    OUTPUT_W_FORMULAE = "output_with_formulae"
    OUTPUT_WO_FORMULAE = "output_without_formulae"

class InputTableHeaders(BaseTableHeaders):
    '''Input table headers in IOSelection.'''

    FILENAME = 0
    COUNTRY = auto()
    WORKSHEET = auto()

class OutputTableHeaders(BaseTableHeaders):
    '''Output table headers in IOSelection.'''

    FILENAME = 0
    WORKSHEET = auto()
    MONTH = auto()
    YEAR = auto()

class SpecialStrings(BaseTableHeaders):
    '''Enum for selecting worksheets.'''

    SELECT_WORKSHEET = "<select worksheet>"
    XLSX = ".xlsx"
    UTF_8 = "utf-8"
    DATA_ONLY_EXCEL = "_wizard_data_only.xlsx"
    ZERO_HOURS = "0.00"
    MISSING = "Missing"
    DEFAULT_PADDING = "5" # Must be cast to int

class OutputFile(BaseTableHeaders):
    '''Enum for original and copy output files.
    The original has formulae preserved (data_only = False).'''

    FIRST_ENTRY = 0
    SECOND_ENTRY = auto()

class ButtonNames(BaseTableHeaders):
    '''Enum of names to display on buttons.'''

    ADD = "Add"
    REMOVE = "Remove"
    SELECT_ALL = "Select all"
    DESELECT_ALL = "Deselect all"

#########################
#### Project ID Enum ####
#########################

class ProjectIDTableHeaders(BaseTableHeaders):
    '''Project ID headers in project selection.'''

    PROJECT_ID = 0
    DESCRIPTION = auto()
    FILENAME = auto()

#################################
#### Employee Selection Enum ####
#################################

class EmployeeTableHeaders(BaseTableHeaders):
    '''Employee table headers in employee selection.'''

    EMPLOYEE = 0
    WORKSHEET = auto()
    COORDINATE = auto()

######################
#### Summary Enum ####
######################

class SummaryTableHeaders(BaseTableHeaders):
    '''Summary table headers in summary selection.'''

    EMPLOYEE = 0
    PREDICTED_HOURS = auto()
    ACCUMULATED_HOURS = auto()
    DEVIATION = auto()
    PROJECT_ID = auto()
    OUTPUT_WORKSHEET = auto()
    COORDINATE = auto()

##################
#### Log Enum ####
##################

class LogTableHeaders(BaseTableHeaders):
    '''Summary table headers in summary selection.'''

    EMPLOYEE = "Employee"
    PREDICTED_HOURS = "Predicted Hours"
    ACCUMULATED_HOURS = "Accumulated Hours"
    DEVIATION = "Deviation"
    PROJECT_ID = "Project ID"
    COORDINATE = "Coordinate"

#######################
#### Coutries Enum ####
#######################

class CountriesEnum(BaseTableHeaders):
    '''Enum of countries.'''

    GERMANY = "Germany"
    ROMANIA = "Romania"

###################
#### Yaml Enum ####
###################

class YamlEnum(BaseTableHeaders):
    '''Enum of top level yaml config entries.'''

    COUNTRIES = "countries"
    DEVIATIONS = "deviations"
    ROW_ANCHORS = "row_anchors"


############################################
#### To be replaced with a yaml handler ####
############################################

NON_NAMES = [
    'MA Name\nStartdatum',
    'GWR REG',
    'Verhandlung',
    'h/Mnt RUM',
    'h/Mnt REG',
    'h/Mnt\ngesamt',
    'Anzahl\nMA'
]

##########################
#### Abstract classes ####
##########################

class YamlHandler(ABC):
    '''Abstract class for deserialising yaml files.'''

    CONFIG_PATH = path.join(path.dirname(__file__), "config_data.yaml")

    def __init__(self):
        super().__init__()
        self.encoding = SpecialStrings.UTF_8.value
        self._load_yaml_data()
    
    @abstractmethod
    def _process_yaml(self,
                      yaml_data: dict) -> None:
        '''Abstract method. Processes the data of the yaml file.'''

        raise NotImplementedError # To override

    def _load_yaml_data(self) -> None:
        '''Template for loading yaml data.'''

        try:
            with open(self.CONFIG_PATH, 'r', encoding=SpecialStrings.UTF_8.value) as yaml_file:
                yaml_data = yaml.safe_load(yaml_file)
                self._process_yaml(yaml_data)
        except FileNotFoundError as exc:
            raise FileNotFoundError("The config file could not be found.") from exc
        except UnicodeDecodeError as exc:
            raise exc

#############################
#### Location management ####
#############################

@dataclass
class FilterHeaders:
    '''Data class for worksheet header strings used for filtering.
    These data are received from LocaleData.'''

    name: str
    proj_id: str # Project number
    description: str # Project's short description
    hours: str
    date: str

@dataclass
class FilePatternData:
    '''Parent data class for establishing the Excel file's naming.'''

    # German or external input timesheets or output budget file
    file_type: str
    # Regular expresion to filter for file in open file dialog.
    file_patterns: list[str]

@dataclass
class InputLocaleData(FilePatternData):
    '''Child data class for establishing the Excel file's locale details.'''

    country: str # Country name
    exp_sheet_name: str # Expected worksheet name
    filter_headers: FilterHeaders = field(default_factory=dict)
    # An extendable list of cities from where the files may originate.
    # This attribute may be removed in a future update.
    cities: list[str] = field(default_factory=list)

    def __post_init__(self):
        '''Init the filter headers from the data received from the country data dataclass.'''

        self.filter_headers = FilterHeaders(**self.filter_headers)

@dataclass
class CountryData(YamlHandler):
    '''Data class using the LocaleData data class to deserialise the yaml config file.'''

    countries: list[InputLocaleData] = field(default_factory=list)

    def __post_init__(self):
        YamlHandler.__init__(self)
        self._load_yaml_data()

    def _process_yaml(self, yaml_data) -> None:
        '''Processes the yaml data.'''

        country_data = yaml_data.get(YamlEnum.COUNTRIES.value, [])
        self.countries = [InputLocaleData(**locale_data) for locale_data in country_data]

    def get_locale_by_country(self, country: str) -> InputLocaleData:
        '''Returns the locale as per given country name.'''

        return next((locale for locale in self.countries if locale.country == country), None)

####################################
#### Hours Deviation management ####
####################################

@dataclass()
class HoursDeviation(YamlHandler):
    '''Data class to define the deviation thresholds for predicted and accumulated hours.'''
    strong_dev: Optional[float] = None
    weak_dev: Optional[float] = None

    def __post_init__(self):
        YamlHandler.__init__(self)
        self._load_yaml_data()

    def _process_yaml(self, yaml_data) -> None:
        '''Processes the yaml data.'''

        self.__dict__.update(yaml_data.get(YamlEnum.DEVIATIONS.value, {}))

#################################
#### Budget Sheet management ####
#################################

@dataclass(eq=False) # Set to false to allow lru caching of futils.locate_employee_range(...)
class EmployeeRowAnchors(YamlHandler):
    '''Data class to define the anchor strings of the row containing the employee names.'''
    start_anchor: str = ""
    end_anchor: str = ""

    def __post_init__(self):
        YamlHandler.__init__(self)
        self._load_yaml_data()

    def _process_yaml(self, yaml_data) -> None:
        '''Processes the yaml data.'''

        self.__dict__.update(yaml_data.get(YamlEnum.ROW_ANCHORS.value, {}))

@dataclass(eq=False) # Set to false to allow lru caching of futils.locate_employee_range(...)
class EmployeeRange:
    '''Data class for the range within which the employee names are located.'''

    start_cell: str = ""
    end_cell: str = ""

    @property
    def start_col_idx(self) -> int:
        '''Get the column integer of the start cell.'''

        return xlutils.coordinate_to_tuple(self.start_cell)[1]

    @property
    def end_col_idx(self) -> int:
        '''Get the column integer of the end cell.'''

        return xlutils.coordinate_to_tuple(self.end_cell)[1]

    @property
    def start_row_idx(self) -> int:
        '''Get the row integer of the start cell.'''

        return xlutils.coordinate_to_tuple(self.start_cell)[0]

    @property
    def end_row_idx(self) -> int:
        '''Get the row integer of the end cell.'''

        return xlutils.coordinate_to_tuple(self.end_cell)[0]

@dataclass
class EmployeeHours:
    '''Data class for managing the predicted and accumulated hours
    per employee.'''

    predicted_hours: Optional[float|int] = None
    pre_hours_colour: Optional[Qt.GlobalColor] = Qt.GlobalColor.black
    accumulated_hours: Optional[float|int] = None
    acc_hours_colour: Qt.GlobalColor = Qt.GlobalColor.black
    hours_coord: Optional[str] = None
    thresholds: HoursDeviation = field(default_factory=HoursDeviation)
    deviation: Optional[str] = None

    def set_deviation(self) -> None:
        '''Sets the deviation by percentage between predicted and accumulated hours.'''

        pre_hrs = self.predicted_hours
        acc_hrs = self.accumulated_hours
        if not isinstance(acc_hrs, (float, int)):
            acc_hrs = 0
        if acc_hrs == 0 and pre_hrs == 0:
            self.deviation = " "
        else:
            # Use only absolute values to ensure correct calculation of 1 - frac
            pre_hrs = abs(pre_hrs)
            acc_hrs = abs(acc_hrs)
            frac = (
                min(acc_hrs, pre_hrs)/
                max(acc_hrs, pre_hrs)
            )
            if 1 - frac >= self.thresholds.strong_dev:
                self.deviation = "Warning! Strong deviation!"
            elif 1 - frac >= self.thresholds.weak_dev:
                self.deviation = "Weak deviation"
            else:
                self.deviation = "Negligible"

@dataclass()
class Employee:
    '''Data class for managing employee name location and related hours
    in the given worksheet. Projects where the employee registered hours
    will be saved here.'''

    name: str
    found_projects: dict[int|str, list[str]] = field(default_factory=dict)
    hours: EmployeeHours = field(default_factory=EmployeeHours)


@dataclass(init=False)
class SelectedDate:
    '''Data class for the located date for which hours will be recorded per employee
    in the budgeting file.'''

    month: Optional[int]
    year: Optional[int]
    row: Optional[int]

##############################
#### Worksheet management ####
##############################

@dataclass
class SelectedSheet:
    '''Data class for worksheet names.'''

    sheet_name: str
    sheet_object: Worksheet

@dataclass
class ManagedWorksheet:
    '''Parent data class for worksheet data.'''

    selected_sheet: Optional[SelectedSheet] = None
    sheet_names: list[str] = field(default_factory=list)


@dataclass
class ManagedInputWorksheet(ManagedWorksheet):
    '''Child data class for input worksheet data.'''

    # The project number is not provided at init/postinit
    # key: project ID (network number [int] or PSP element [str]),
    # value: list of short project desciptions
    selectable_project_ids: dict[str, list[str]] = field(default_factory=dict)
    # Dictionary of project IDs (network number [int] or PSP element [str]): descriptions
    selected_project_ids: dict[str, list[str]] = field(default_factory=dict)
    indexed_headers: dict[str, int] = field(default_factory=dict)

    def index_headers(self) -> None:
        '''Retrieves the table headers from the given worksheet
        and maps them to their column indices.'''

        # sheet object [1] is row one, where the headers are
        for idx, cell in enumerate(self.selected_sheet.sheet_object[1]):
            if isinstance(cell.value, str):
                self.indexed_headers[cell.value] = idx

    def yield_from_project_id_and_desc(self) -> Iterator[tuple[str, list[str]]]:
        '''Yields from the project ID and description,
        one at a time in a tuple.'''

        yield from self.selectable_project_ids.items()

    def set_selectable_project_ids(self, proj_id: str, proj_desc: str, name: str) -> None:
        '''Extracts all project IDs in the selected worksheet and saves the 
        data in the selectable project IDs dictionary.'''

        # Get the worksheet object
        sheet_object = self.selected_sheet.sheet_object
        # Get the project ID column index
        project_id_col = self.indexed_headers.get(proj_id)
        # Get the project description column index
        project_desc_col = self.indexed_headers.get(proj_desc)
        # Get the project person column index
        person_col = self.indexed_headers.get(name)
        for row in sheet_object.iter_rows(min_row=2):
            # Get the contents of each row under the given headers
            id_value = row[project_id_col].value
            desc_value = row[project_desc_col].value
            person_value = row[person_col].value
            # Only add the project ID and description if they exist alongside an employee
            if id_value and desc_value and person_value:
                id_value = str(id_value)
                desc_value = str(desc_value)
                person_value = str(person_value)
                if id_value not in self.selectable_project_ids:
                    # If the ID is not in the dictionary, make a new list for descriptions
                    self.selectable_project_ids[id_value] = []
                if desc_value not in self.selectable_project_ids[id_value]:
                    # Append the descriptions per ID, if it has not already been appended
                    self.selectable_project_ids[id_value].append(desc_value)

@dataclass
class ManagedOutputWorksheet(ManagedWorksheet):
    '''Child data class for output worksheet data.'''

    # Replaced through IOSelection
    selected_date: SelectedDate = field(default_factory=SelectedDate)
    employee_row_anchors: EmployeeRowAnchors = field(default_factory=EmployeeRowAnchors)
    # Start cell: coord, end cell: coord
    employee_range: EmployeeRange = field(default_factory=EmployeeRange)
    # Key: Cell coord; value: Employee object
    selected_employees: dict[str, Employee] = field(default_factory=dict)

    def set_sheet_names(self, sheet_names) -> None:
        '''Set sheet names.'''

        self.sheet_names = sheet_names

    def set_selected_sheet(self, sheet_name: str, sheet_object: Worksheet) -> None:
        '''Set selected sheet.'''

        # Save the worksheet data
        self.selected_sheet = SelectedSheet(
            sheet_name,
            sheet_object
            )
        self.selected_sheet.sheet_name = sheet_name
        self.selected_sheet.sheet_object = sheet_object
        # Check whether the employees are located in the worksheet
        self.employee_range = EmployeeRange()
        futils.locate_employee_range(
            self.selected_sheet.sheet_object,
            self.employee_range,
            self.employee_row_anchors
        )

    def set_budgeting_date(self,
                           file_path: str,
                           sheet_name: str,
                           selected_month: str,
                           selected_year: str) -> None:
        '''Sets the budgeting date with the row it is located in the worksheet.'''

        # Convert dates to integers and put in a tuple
        month_year = (MONATE_KURZ_DE.get(selected_month), int(selected_year))
        budgeting_dates = futils.get_budgeting_dates(file_path, sheet_name)
        for tup in budgeting_dates:
            if tup[:2] == month_year:
                month, year, row = tup
                break
        else:
            raise BudgetingDatesNotFound(
                selected_month,
                selected_year,
                sheet_name,
                path.basename(file_path))
        self.selected_date.month = month
        self.selected_date.year = year
        self.selected_date.row = row

    def set_selected_employees(self, coord_name: list[tuple[str, str]]) -> None:
        '''Save the coordinate in the worksheet with the selected employee.'''

        for coord, name in coord_name:
            self.selected_employees[coord] = Employee(name)

    def set_predicted_hours(self, coord_hours: dict[str, float|int]) -> None:
        '''Sets the attribute with the predicted hours found in the worksheet.'''

        for coord, hours in coord_hours.items():
            employee = self.selected_employees.get(coord)
            if employee:
                if not isinstance(hours, (float, int)):
                    hours = 0.0
                employee.hours.predicted_hours = hours

    def set_predicted_hours_colour(self) -> None:
        '''Set the color formatting of the employee's predicted hours.
        Predicted hours may be coloured as the user wishes but recorded
        hours must be set to black (default, theme or RGB).'''

        for emp in self.yield_from_selected_employee():
            if emp.hours.hours_coord:
                # Check each employee's predicted hours colour
                font_colour = self.selected_sheet.sheet_object[emp.hours.hours_coord].font.color
                if not font_colour:
                    # Default (black): Already recorded.
                    # Show red to get the user's attention in summary
                    emp.hours.pre_hours_colour = Qt.GlobalColor.red
                elif (
                    font_colour.type == 'theme' and
                    font_colour.theme == 1 and
                    font_colour.tint == 0.0
                ):
                    # Black theme: Already recorded.
                    # Show red to get the user's attention in summary
                    emp.hours.pre_hours_colour = Qt.GlobalColor.red
                elif (
                    font_colour.type == 'rgb' and
                    font_colour.rgb == 'FF000000'
                ):
                    # Black RGB: Already recorded.
                    # Show red to get the user's attention in summary
                    emp.hours.pre_hours_colour = Qt.GlobalColor.red

    def clear_predicted_hours(self) -> None:
        '''Clears the recorded employee names and respective hours.'''

        self.selected_employees.clear()

    def yield_from_selected_employee(self) -> Iterator[Employee]:
        '''Yields from the selected employee.'''

        yield from self.selected_employees.values()

#############################
#### Workbook management ####
#############################

@dataclass
class ManagedWorkbook:
    '''Parent data class for workbook data.'''

    file_path: str # Mandatory parameter
    file_name: Optional[str] = None
    data_only: bool = False # By default keep the formulae in the workbook
    workbook_object: Optional[Workbook] = None

    def __post_init__(self):
        self.file_name = path.basename(self.file_path)

    def __eq__(self, other):
        '''Compare data objects by file names and data_only.'''

        if isinstance(other, ManagedWorkbook):
            return self.file_name == other.file_name and self.data_only == other.data_only
        return False

@dataclass
class ManagedOutputWorkbook(ManagedWorkbook):
    '''Child data class for output workbook data (budget).'''

    managed_sheet_object: Optional[ManagedOutputWorksheet] = None

    def __post_init__(self):
        super().__post_init__()
        self.workbook_object = futils.try_load_workbook(self, ManagedOutputWorkbook, self.data_only)

    def init_output_worksheet(self) -> None:
        '''Init output worksheet.'''

        self.managed_sheet_object = ManagedOutputWorksheet()
        self.managed_sheet_object.set_sheet_names(self.workbook_object.sheetnames)

    def save_output_workbook(self) -> None:
        '''Saves the workbook with its given file path.'''

        self.workbook_object.save(self.file_path)

@dataclass
class ManagedInputWorkbook(ManagedWorkbook):
    '''Child data class for input workbook data (timesheets).
    The data is provided from a deep copy made from the Location
    Manager's retrieved data.'''

    locale_data: Optional[InputLocaleData] = None
    managed_sheet_object: Optional[ManagedInputWorksheet] = None

    def __post_init__(self):
        super().__post_init__()
        self.workbook_object = futils.try_load_workbook(self, ManagedOutputWorkbook)

    def set_locale_data(self, country_data: CountryData, country_name: str) -> None:
        '''Sets the locale data.'''
        
        # A deep copy is not required because the data object will not change
        self.locale_data = next(
            (locale for locale in country_data.countries
            if locale.country == country_name),
            None)

    def init_input_worksheet(self):
        '''Init input worksheet.'''

        if len(self.workbook_object.sheetnames) <= 1:
            sheet_name = self.workbook_object.sheetnames[0]
        else:
            sheet_name = self.locale_data.exp_sheet_name
        self.managed_sheet_object = ManagedInputWorksheet(
            SelectedSheet(
                sheet_name,
                self.workbook_object[sheet_name]
            )
        )
        self.managed_sheet_object.index_headers()

@dataclass
class WorkbookManager:
    '''Data class for tracking workbooks.'''

    workbooks: list[ManagedInputWorkbook | ManagedOutputWorkbook] = field(default_factory=list)

    def try_add_workbook(self,
                         workbook: ManagedInputWorkbook | ManagedOutputWorkbook) -> None:
        '''Try to add the workbook to tracking.'''

        retrieved_wb = self.get_workbook_by_name(workbook.file_name)
        if retrieved_wb != workbook:
            self.workbooks.append(workbook)
        else:
            raise WorkbookAlreadyTracked(workbook.file_name)

    def add_input_workbook(self, file_path: str) -> None:
        '''Add input workbooks by filename.'''

        workbook = ManagedInputWorkbook(file_path)
        self.try_add_workbook(workbook)

    def add_output_workbook(self, file_path: str, data_only: bool = False) -> None:
        '''Add input workbooks by filename.'''

        workbook = ManagedOutputWorkbook(file_path=file_path, data_only=data_only)
        self.try_add_workbook(workbook)

    def get_workbook_by_name(self,
                     file_name: str,
    ) -> ManagedInputWorkbook | ManagedOutputWorkbook | None:
        '''Retrieves the first workbook in the list by file name.'''

        return next((wb for wb in self.workbooks
                     if file_name == wb.file_name),
                     None)

    def yield_workbooks_by_type(self,
                     wb_class: ManagedInputWorkbook | ManagedOutputWorkbook,
    ) -> Iterator[ManagedInputWorkbook | ManagedOutputWorkbook]:
        '''Returns a generator of workbooks in the list by type.'''

        for wb in self.workbooks:
            if isinstance(wb, wb_class):
                yield wb

    def remove_workbook(self,
                        file_name: str) -> None:
        '''Add input/output workbooks by filename.'''

        if file_name:
            wb = self.get_workbook_by_name(file_name)
            # Delete workbook object if it was created
            if wb:
                self.workbooks.remove(wb)
                del wb

########################
#### Log management ####
########################

@dataclass
class FileMetaData:
    '''Encapsulates all necessary log data.'''

    log_file_path: str
    selected_date: datetime
    input_workbooks: list[str]
    output_file_name: str
    output_worksheet_name: str

@dataclass
class TableStructure:
    '''Encapsulates table-related metadata.'''

    headers: list[str]
    col_widths: list[int]
