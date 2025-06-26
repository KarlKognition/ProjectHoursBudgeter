'''
Package
-------
Data Handling

Module Name
---------
Worksheet Management

Author
-------
Karl Goran Antony Zuvela

Description
-----------
Provides worksheet management.
'''
from dataclasses import dataclass, field
from typing import Optional, Iterator, TYPE_CHECKING
from openpyxl.worksheet.worksheet import Worksheet
from PyQt6.QtCore import Qt
#           --- First party libraries ---
import phb_app.data.selected_date as sd
import phb_app.data.employee_management as emp
import phb_app.templating.types as t
import phb_app.utils.employee_utils as eu

if TYPE_CHECKING:
    import phb_app.data.workbook_management as wm

#           --- DATA CONTAINERS ---

@dataclass
class SelectedSheet:
    '''Data class for worksheet names.'''
    sheet_name: str
    sheet_object: Worksheet

@dataclass
class InputWorksheetContext:
    '''Data class for input worksheet data.'''
    selected_sheet: Optional[SelectedSheet]
    sheet_names: list[str] = field(default_factory=list)
    selectable_project_ids: t.ProjectsDict = field(default_factory=dict)
    selected_project_ids: t.ProjectsDict = field(default_factory=dict)
    indexed_headers: dict[str, int] = field(default_factory=dict)

@dataclass
class OutputWorksheetContext:
    '''Data class for output worksheet data.'''
    selected_sheet: Optional[SelectedSheet] = None
    sheet_names: list[str] = field(default_factory=list)
    selected_date: sd.SelectedDate = field(default_factory=sd.SelectedDate)
    employee_row_anchors: emp.EmployeeRowAnchors = field(default_factory=emp.EmployeeRowAnchors)
    employee_range: emp.EmployeeRange = field(default_factory=emp.EmployeeRange)
    selected_employees: dict[str, emp.Employee] = field(default_factory=dict)

#           --- SERVICE CLASSES ---

class InputWorksheetService:
    '''Service class for managing an input worksheet.'''
    def __init__(self, worksheet: InputWorksheetContext):
        self.worksheet = worksheet

    def set_sheet_names(self, sheet_names: list[str]) -> None:
        '''Set the sheet names in the output worksheet data.'''
        self.worksheet.sheet_names = sheet_names

    def index_headers(self) -> None:
        '''Indexes the headers in the selected worksheet.'''
        if self.worksheet.selected_sheet:
            for idx, cell in enumerate(self.worksheet.selected_sheet.sheet_object[1]):
                if isinstance(cell.value, str):
                    self.worksheet.indexed_headers[cell.value] = idx

    def yield_project_id_and_desc(self) -> Iterator[t.ProjectsTup]:
        '''Yields from the project ID and description, one at a time in a tuple.'''
        yield from self.worksheet.selectable_project_ids.items()

    def set_selectable_project_ids(self, proj_id: str, proj_desc: str, name: str) -> None:
        '''Extracts all project IDs in the selected worksheet and saves the 
        data in the selectable project IDs dictionary.'''
        # Return early if no worksheet is selected
        if not self.worksheet.selected_sheet:
            return
        # Get the worksheet object
        sheet_object = self.worksheet.selected_sheet.sheet_object
        # Get the column indices for project ID, description, and person
        project_id_col = self.worksheet.indexed_headers.get(proj_id)
        project_desc_col = self.worksheet.indexed_headers.get(proj_desc)
        person_col = self.worksheet.indexed_headers.get(name)
        # Iterate over each row in the worksheet, skipping the header
        for row in sheet_object.iter_rows(min_row=2):
            # Extract values for project ID, description, and person from the row
            id_value = row[project_id_col].value
            desc_value = row[project_desc_col].value
            person_value = row[person_col].value
            # Only process rows where all three values are present
            if id_value and desc_value and person_value:
                # Convert values to strings for consistency
                id_value = str(id_value)
                desc_value = str(desc_value)
                person_value = str(person_value)
                # If this project ID is not yet in the dictionary, add it with an empty list
                if id_value not in self.worksheet.selectable_project_ids:
                    self.worksheet.selectable_project_ids[id_value] = []
                # If this description is not already associated with the project ID, append it
                if desc_value not in self.worksheet.selectable_project_ids[id_value]:
                    self.worksheet.selectable_project_ids[id_value].append(desc_value)

class OutputWorksheetService:
    '''Service class for managing an output worksheet.'''
    def __init__(self, worksheet: OutputWorksheetContext):
        self.worksheet = worksheet

    def set_sheet_names(self, sheet_names: list[str]) -> None:
        '''Set the sheet names in the output worksheet data.'''
        self.worksheet.sheet_names = sheet_names

    def set_selected_sheet(self, wb_ctx: "wm.OutputWorkbookContext", sheet_name: str) -> None:
        '''Set selected sheet.'''
        # Save the worksheet data
        self.worksheet.selected_sheet = SelectedSheet(sheet_name, wb_ctx.mngd_wb.workbook_object[sheet_name])

    def compute_employee_range(self) -> None:
        '''Create the employee range.'''
        self.worksheet.employee_range = emp.EmployeeRange()
        eu.set_employee_range(self.worksheet.selected_sheet.sheet_object, self.worksheet.employee_range, self.worksheet.employee_row_anchors)

    def set_selected_employees(self, coord_name: list[tuple[str, str]]) -> None:
        '''Save the coordinate in the worksheet with the selected employee.'''
        for coord, name in coord_name:
            self.worksheet.selected_employees[coord] = emp.Employee(name)

    def set_predicted_hours(self, coord_hours: dict[str, float|int]) -> None:
        '''Sets the predicted hours for each employee based on their coordinates.'''
        for coord, hours in coord_hours.items():
            employee = self.worksheet.selected_employees.get(coord)
            if employee:
                pred_hrs = float(hours) if isinstance(hours, (float, int)) else 0.0
                employee.hours.predicted_hours = pred_hrs

    def set_predicted_hours_colour(self) -> None:
        '''Set the color formatting of the employee's predicted hours.'''
        for employee in self.yield_from_selected_employee():
            if employee.hours.hours_coord:
                # Check each employee's predicted hours colour
                sheet_obj = self.worksheet.selected_sheet.sheet_object
                font_colour = sheet_obj[employee.hours.hours_coord].font.color
                if not font_colour or (
                    font_colour.type == 'theme' and
                    font_colour.theme == 1 and
                    font_colour.tint == 0.0
                ) or (
                    font_colour.type == 'rgb' and
                    font_colour.rgb == 'FF000000'
                ):
                    # Default (black): Already recorded.
                    # Show red to get the user's attention in summary
                    employee.hours.pre_hours_colour = Qt.GlobalColor.red

    def clear_predicted_hours(self) -> None:
        '''Clears the recorded employee names and respective hours.'''
        self.worksheet.selected_employees.clear()

    def yield_from_selected_employee(self) -> Iterator[emp.Employee]:
        '''Yields from the selected employees.'''
        yield from self.worksheet.selected_employees.values()

#           --- MODULE FACTORY FUNCTIONS ---

def _create_input_worksheet_context(
    sheet_name: str,
    sheet_object: Worksheet
) -> InputWorksheetContext:
    '''Public module level. Creates an InputWorksheet for the given sheet name and sheet object.'''
    selected = SelectedSheet(sheet_name=sheet_name, sheet_object=sheet_object)
    return InputWorksheetContext(selected_sheet=selected)

def init_output_worksheet(context: "wm.OutputWorkbookContext") -> None:
    """Public module level. Init output worksheet."""
    worksheet = _create_output_worksheet_context()
    service = OutputWorksheetService(worksheet=worksheet)
    service.set_sheet_names(context.mngd_wb.workbook_object.sheetnames)
    context.managed_sheet = worksheet
    context.worksheet_service = service

def _create_output_worksheet_context() -> OutputWorksheetContext:
    '''Public module level. Creates an empty OutputWorksheet, ready for sheet selection in the UI.'''
    return OutputWorksheetContext()

def init_input_worksheet(context: "wm.InputWorkbookContext") -> None:
    """Public module level. Init input worksheet."""
    sheetnames = context.mngd_wb.workbook_object.sheetnames
    # If there is only one sheet, use that;
    # otherwise, use the locale data's expected sheet name.
    sheet_name = (
        sheetnames[0]
        if len(sheetnames) <= 1
        else context.locale_data.exp_sheet_name
    )
    sheet_obj = context.mngd_wb.workbook_object[sheet_name]
    managed_sheet = _create_input_worksheet_context(sheet_name, sheet_obj)
    service = InputWorksheetService(worksheet=managed_sheet)
    service.set_sheet_names(context.mngd_wb.workbook_object.sheetnames)
    context.managed_sheet = managed_sheet
    context.worksheet_service = service
    service.index_headers()
